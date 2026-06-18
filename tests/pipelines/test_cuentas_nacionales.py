from pathlib import Path

import openpyxl
import pandas as pd

from sipsa_leche.pipelines.cuentas_nacionales.nodes import calcular_excluidas, generar_leche_cruda


def _crear_base_leche_cruda(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LECHE CRUDA"
    ws.append(
        [
            "Anno",
            "Mes",
            "Trimestre",
            "Producido_mes_DANE_Recolectado  ",
            "Factor_Correccion_mes",
            "Producido_mes_DANE  ",
            "Variacion_DANE",
            "Producido_mes_NAL_11",
            "Producido_mes_NAL_12",
            "Producido_mes_NAL_13",
        ]
    )
    historico = [
        (2025, "Marzo", "I"),
        (2025, "Abril", "II"),
        (2025, "Mayo", "II"),
        (2025, "Junio", "II"),
        (2025, "Julio", "III"),
        (2025, "Agosto", "III"),
        (2025, "Septiembre", "III"),
        (2025, "Octubre", "IV"),
        (2025, "Noviembre", "IV"),
        (2025, "Diciembre", "IV"),
        (2026, "Enero", "I"),
        (2026, "Febrero", "I"),
        (2026, "Marzo", "I"),
    ]
    for row_idx, row in enumerate(historico, start=2):
        ws.cell(row_idx, 1).value = row[0]
        ws.cell(row_idx, 2).value = row[1]
        ws.cell(row_idx, 3).value = row[2]
        ws.cell(row_idx, 4).value = 1
        ws.cell(row_idx, 5).value = "=28/(7*4)"
        ws.cell(row_idx, 6).value = f"=D{row_idx}*E{row_idx}"
        ws.cell(row_idx, 7).value = f"=(F{row_idx}/F{max(row_idx - 1, 2)})-1"
        ws.cell(row_idx, 8).value = f"=((F{row_idx}-F2)/F2)*100"
        ws.cell(row_idx, 10).value = f"=(G{row_idx}+1)*J{max(row_idx - 1, 2)}"

    # Hoja LecheDANE: espejo del histórico con valores D = 1.0 por fila
    ws_ld = wb.create_sheet("LecheDANE")
    ws_ld.cell(1, 1).value = "Anno"
    ws_ld.cell(1, 2).value = "Mes"
    ws_ld.cell(1, 3).value = "Trimestre"
    ws_ld.cell(1, 4).value = "D"
    for row_idx, row in enumerate(historico, start=2):
        ws_ld.cell(row_idx, 1).value = row[0]
        ws_ld.cell(row_idx, 2).value = row[1]
        ws_ld.cell(row_idx, 3).value = row[2]
        ws_ld.cell(row_idx, 4).value = 1.0

    # Hoja trimes: una fila por trimestre con año y código de trimestre
    ws_tr = wb.create_sheet("trimes")
    trimestres = [
        (2025, "I"), (2025, "II"), (2025, "III"), (2025, "IV"),
        (2026, "I"),
    ]
    for r_idx, (anno, trim) in enumerate(trimestres, start=1):
        ws_tr.cell(r_idx, 1).value = anno
        ws_tr.cell(r_idx, 2).value = trim
        ws_tr.cell(r_idx, 4).value = 1.0

    wb.save(path)
    wb.close()


def test_generar_leche_cruda_usa_total_macro_en_columna_d(tmp_path, monkeypatch):
    base_path = tmp_path / "LECHE_CRUDA_EST_BASE.xlsx"
    _crear_base_leche_cruda(base_path)
    monkeypatch.chdir(tmp_path)

    variacion_macro = pd.DataFrame(
        {
            "T_PRODUCCION_MACROMAR": [
                10_757_541,
                1_648_142,
                2_473_881,
                1_477_676,
                4_374_475,
            ]
        }
    )

    resumen = generar_leche_cruda(
        variacion_macro=variacion_macro,
        total_excluidas=123,
        n_semanas=4,
        ruta_leche_cruda_base=str(base_path),
        periodo="032026",
        mes_actual="MAR",
    )

    output_path = tmp_path / "data" / "08_reporting" / "LECHE_CRUDA_032026.xlsx"
    wb = openpyxl.load_workbook(output_path, data_only=False)
    ws = wb["LECHE CRUDA"]

    assert ws.max_row == 14
    assert ws.cell(14, 4).value == "=20731715-123"
    assert ws.cell(14, 5).value == "=31/(7*4)"
    assert ws.cell(14, 6).value == "=D14*E14"
    assert ws.cell(14, 7).value == "=(F14/F13)-1"
    assert ws.cell(14, 8).value == "=((F14-F2)/F2)*100"
    assert ws.cell(14, 10).value == "=(G14+1)*J13"
    wb.close()

    assert resumen["D_produccion_recolectada"] == 20_731_715 - 123
    assert resumen["total_excluidas"] == 123


def _crear_base_excluidas(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Hoja1"
    wb.create_sheet("Hoja2")

    ws1.cell(2, 1).value = "DEPARTAMENTO"
    ws1.cell(2, 2).value = "IDFINCA"
    ws1.cell(2, 11).value = "PRODC032022"
    ws1.cell(2, 58).value = "PRODC022026"
    ws1.cell(2, 59).value = "PRODC032026"
    ws1.cell(3, 1).value = "Antioquia"
    ws1.cell(3, 2).value = "123"
    ws1.cell(4, 1).value = "Total"
    ws1.cell(5, 1).value = "Variacion"
    ws1.cell(39, 3).value = "salen"
    ws1.cell(40, 3).value = "entran"
    ws1.cell(41, 3).value = "diferencia"

    wb.save(path)
    wb.close()


def test_calcular_excluidas_escribe_marzo_2026_en_bg_sin_tocar_marzo_2022(
    tmp_path,
    monkeypatch,
):
    base_path = tmp_path / "Excluidas_leche.xlsx"
    _crear_base_excluidas(base_path)
    monkeypatch.chdir(tmp_path)

    variacion_finca = pd.DataFrame(
        {
            "IDFINCA": ["0000123", "0000456", "0000789"],
            "T_PROD_FEB": [900, None, 700],
            "T_PROD_MAR": [1200, 500, None],
        }
    )

    total = calcular_excluidas(
        variacion_finca=variacion_finca,
        ruta_excluidas=str(base_path),
        periodo="032026",
        mes_actual="MAR",
        mes_anterior="FEB",
    )

    output_path = tmp_path / "data" / "08_reporting" / "Excluidas_leche_032026.xlsx"
    wb = openpyxl.load_workbook(output_path)
    ws1 = wb["Hoja1"]

    assert total == 1200
    assert ws1["K2"].value == "PRODC032022"
    assert ws1["BG2"].value == "PRODC032026"
    assert ws1["BG3"].value == 1200
    assert ws1["BG4"].value == 1200
    assert ws1["BG39"].value == 700
    assert ws1["BG40"].value == 500
    assert ws1["BG41"].value == "=+BG40-BG39"

    wb.close()
