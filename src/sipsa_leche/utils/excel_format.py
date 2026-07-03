"""Formato de publicación para los cuadros de salida mensuales.

Replica el formato visual de los cuadros generados por el proceso SAS original
(fuente, colores de relleno, bordes y formato numérico por columna), reverso-
ingenierizado a partir de CUADROS_032026_salida.xlsx / _TOT_salida.xlsx.

Transforma los Excel generados por pandas en el formato de entrega:

  CUADROS_{PERI}.xlsx (pub):
    - FINCA:        anchos de columna, encabezados sin negrita, hoja oculta
    - MUNICIPIO
    - DEPARTAMENTO  →  2 filas de encabezado (mes anterior / mes actual /
    - MACROREGION      variación / tendencia); encabezados en fuente MS Sans
                        Serif 10pt con relleno gris claro y bordes finos;
                        columna "Precio promedio" en rojo negrita con relleno
                        rosado; formato numérico específico por columna.

  CUADROS_{PERI}_TOT.xlsx:
    - Todas las hojas: encabezados sin negrita + anchos de columna (sin
      relleno/bordes adicionales — el archivo TOT de SAS tampoco los tiene)
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_BOLD = Font(bold=True)
_NOT_BOLD = Font(bold=False)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_ALIGN = Alignment(vertical="center")

# Encabezados: MS Sans Serif 10pt negrita, relleno gris claro (blanco −5%
# de luminosidad, tema Office 2007-2010 = F2F2F2), bordes finos en las 4 caras.
_HEADER_FONT = Font(name="MS Sans Serif", size=10, bold=True)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")
_THIN = Side(style="thin")
_HEADER_BORDER = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)

# Datos: borde fino izquierda/derecha únicamente (sin horizontal).
_DATA_BORDER = Border(left=_THIN, right=_THIN)

# Columna "Precio promedio"/"PRECIO" (VPRE): rojo negrita MS Sans Serif con
# relleno rosado (tema Accent2 = C0504D al 60% de tinte = E6B9B8) como base;
# formato condicional de Excel "Relleno amarillo con texto amarillo oscuro"
# la sobrescribe cuando el valor está entre -5% y +5% (rango "estable").
_PRECIO_FONT = Font(name="MS Sans Serif", size=10, bold=True, color="FFFF0000")
_PRECIO_FILL = PatternFill(fill_type="solid", fgColor="E6B9B8")
_PRECIO_CF_FONT = Font(color="9C6500")
_PRECIO_CF_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

_FMT_DECIMAL = "0.00"
_FMT_PCT = "0.00%"
_FMT_ACCOUNTING = '_-* #,##0.00_-;\\-* #,##0.00_-;_-* "-"??_-;_-@_-'

# Iniciales → nombre del mes en español (mayúsculas) para encabezado de grupo
_INI_A_NOMBRE: dict[str, str] = {
    "ENE": "ENERO",      "FEB": "FEBRERO",    "MAR": "MARZO",
    "ABR": "ABRIL",      "MAY": "MAYO",       "JUN": "JUNIO",
    "JUL": "JULIO",      "AGO": "AGOSTO",     "SEP": "SEPTIEMBRE",
    "OCT": "OCTUBRE",    "NOV": "NOVIEMBRE",  "DIC": "DICIEMBRE",
}


# ─── Configuraciones por hoja (formato pub) ────────────────────────────────────

def _cfg_municipio(ant: str, act: str) -> dict:
    """Estructura de encabezados para la hoja MUNICIPIO (20 columnas A-T)."""
    nom_ant = _INI_A_NOMBRE.get(ant, ant)
    nom_act = _INI_A_NOMBRE.get(act, act)
    return {
        "grupos": [          # (col_ini, col_fin, label) — 1-indexed
            (6,  11, nom_ant),
            (12, 17, nom_act),
            (18, 19, f"VARIACION {ant}-{act}"),
        ],
        "tendencia_col": 20,   # columna T → merge T1:T2
        "row2_labels": {       # col_1based → etiqueta
            1: "DEPARTAMENTO", 2: "MUNICIPIO",  3: "COD DEP",
            4: "COD MUNI",     5: "IDDEPMUNI",
            6: "MINIMO",       7: "MAXIMO",     8: "PROMEDIO ",
            9: "CV",          10: "NAL %",      11: "DEP %",
           12: "MINIMO",      13: "MAXIMO ",   14: "POMEDIO",
           15: "CV",          16: "NAL %",      17: "DEP %",
           18: "Precio promedio",  19: "Produccion",
        },
        "precio_var_col": 18,   # columna R — negrita en datos
        "row1_height": 31.5,
        "row2_height": 25.5,
        "col_widths": {
            "A": 13.0, "B": 10.0, "C":  8.0, "D":  9.0, "E": 16.0,
            "F": 19.0, "J": 17.0, "K": 15.0, "L": 19.0,
            "P": 17.0, "Q": 15.0, "R": 13.0, "T": 17.0,
        },
        "numfmts": {           # col_1based → formato numérico de datos
            8: _FMT_DECIMAL,   9: _FMT_PCT, 10: _FMT_PCT, 11: _FMT_PCT,
           14: _FMT_DECIMAL,  15: _FMT_PCT, 16: _FMT_PCT, 17: _FMT_PCT,
           18: _FMT_PCT,      19: _FMT_PCT,
        },
    }


def _cfg_departamento(ant: str, act: str) -> dict:
    """Estructura de encabezados para la hoja DEPARTAMENTO (17 columnas A-Q)."""
    nom_ant = _INI_A_NOMBRE.get(ant, ant)
    nom_act = _INI_A_NOMBRE.get(act, act)
    return {
        "grupos": [
            ( 3,  8, nom_ant),
            ( 9, 14, nom_act),
            (15, 16, f"VARIACION {ant}-{act}"),
        ],
        "tendencia_col": 17,   # Q
        "row2_labels": {
            1: "DEPARTAMENTO", 2: "CODIGO ",
            3: "MÍNIMO",       4: "MÁXIMO ",    5: "PROMEDIO",
            6: "SD",           7: "CV",          8: "NAL %",
            9: "MÍNIMO",      10: "MÁXIMO ",    11: "PROMEDIO ",
           12: "SD",          13: "CV",          14: "NAL %",
           15: "Precio promedio",  16: "Produccion",
        },
        "precio_var_col": 15,   # O
        "row1_height": 33.75,
        "row2_height": 25.5,
        "col_widths": {
            "A": 13.0, "B":  8.0, "C": 18.0,
            "F": 17.0, "G": 18.0, "H": 13.0,
            "I": 18.0, "L": 17.0, "M": 18.0, "N": 13.0, "Q": 17.0,
        },
        "numfmts": {
            5: _FMT_DECIMAL,  6: _FMT_DECIMAL,  7: _FMT_PCT,  8: _FMT_ACCOUNTING,
           11: _FMT_DECIMAL, 12: _FMT_DECIMAL, 13: _FMT_PCT, 14: _FMT_ACCOUNTING,
           15: _FMT_PCT,     16: _FMT_PCT,
        },
    }


def _cfg_macroregion(ant: str, act: str) -> dict:
    """Estructura de encabezados para la hoja MACROREGION (16 columnas A-P)."""
    nom_ant = _INI_A_NOMBRE.get(ant, ant)
    nom_act = _INI_A_NOMBRE.get(act, act)
    return {
        "grupos": [
            ( 2,  7, nom_ant),
            ( 8, 13, nom_act),
            (14, 15, f"VARIACION {ant}-{act}"),
        ],
        "tendencia_col": 16,   # P
        "row2_labels": {
            1:  "MACRO",
            2:  "MÍNIMO",      3: "MÁXIMO ",     4: "PROMEDIO ",
            5:  "SD",          6: "CV",           7: "NAL%",
            8:  "MÍNIMO",      9: "MÁXIMO ",     10: "PROMEDIO",
           11:  "SD",         12: "CV",           13: "NAL % ",
           14:  "PRECIO",     15: "PRODUCCION",
        },
        "precio_var_col": 14,   # N
        "row1_height": 51.0,
        "row2_height": 25.5,
        "col_widths": {
            "A": 11.0, "B": 19.0, "G": 16.0,
            "H": 19.0, "M": 16.0, "N": 13.0, "P": 17.0,
        },
        "numfmts": {
            4: _FMT_DECIMAL,  5: _FMT_DECIMAL,  6: _FMT_PCT,  7: _FMT_PCT,
           10: _FMT_DECIMAL, 11: _FMT_DECIMAL, 12: _FMT_PCT, 13: _FMT_PCT,
           14: _FMT_PCT,     15: _FMT_PCT,
        },
    }


# ─── Anchos de columna ─────────────────────────────────────────────────────────

_FINCA_WIDTHS = {
    "A": 13.0, "B": 10.0, "C": 13.0, "D": 8.0,  "E": 9.0,  "F": 8.0,
    "G": 28.0, "H": 13.0, "K": 15.0, "L": 14.0, "M": 15.0, "N": 14.0,
    "O": 13.0, "S": 15.0, "T": 14.0, "U": 15.0, "V": 14.0, "W": 13.0,
    "Y": 17.0, "Z": 13.0,
}

_TOT_WIDTHS: dict[str, dict[str, float]] = {
    "FINCA": _FINCA_WIDTHS,
    "MUNICIPIO": {
        "A": 13.0, "B": 10.0, "C":  8.0, "D":  9.0, "E": 16.0, "F": 19.0,
        "K": 17.0, "L": 23.0, "M": 22.0, "N": 23.0, "O": 17.0, "Q": 13.0,
        "R": 15.0, "S": 19.0, "X": 17.0, "Y": 23.0, "Z": 22.0, "AA": 23.0,
        "AB": 17.0, "AD": 13.0, "AE": 15.0, "AF": 13.0, "AG": 17.0, "AH": 13.0,
    },
    "DEPARTAMENTO": {
        "A": 13.0, "B":  8.0, "C": 18.0, "F": 17.0, "G": 18.0, "H": 14.0,
        "I": 15.0, "K": 16.0, "L": 15.0, "N": 16.0, "P": 13.0, "Q": 18.0,
        "T": 17.0, "U": 18.0, "V": 14.0, "W": 15.0, "Y": 16.0, "Z": 15.0,
        "AB": 16.0, "AD": 13.0, "AF": 17.0, "AG": 13.0,
    },
    "MACROREGION": {
        "A": 11.0, "B": 19.0, "G": 17.0, "H": 23.0, "I": 22.0, "J": 23.0,
        "K": 17.0, "L": 16.0, "M": 19.0, "R": 17.0, "S": 23.0, "T": 22.0,
        "U": 23.0, "V": 17.0, "W": 16.0, "X": 13.0, "Y": 17.0, "Z": 13.0,
    },
}


# ─── Aplicar formato de 2 filas a una hoja ────────────────────────────────────

def _aplicar_2row(ws, cfg: dict) -> None:
    """Convierte una hoja de 1-fila-encabezado en formato de 2 filas con grupos.

    Replica el estilo del cuadro SAS original: encabezado en MS Sans Serif 10pt
    negrita con relleno gris claro y bordes finos (incluye las celdas de
    relleno sin etiqueta, ej. columnas de texto sin grupo); datos con borde
    fino izquierda/derecha, formato numérico por columna, y la columna de
    variación de precio en rojo negrita con relleno rosado.
    """
    ws.insert_rows(1)
    ws.row_dimensions[1].height = cfg["row1_height"]
    ws.row_dimensions[2].height = cfg["row2_height"]
    t_col = cfg["tendencia_col"]

    # Estilo base de encabezado en toda la fila 1 y 2 (incluye celdas sin
    # etiqueta bajo columnas de texto sin grupo, que en SAS también llevan
    # el relleno/borde aunque queden vacías).
    for r in (1, 2):
        for c in range(1, t_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.border = _HEADER_BORDER
            cell.alignment = _CENTER

    # Fila 1 — grupos horizontales (merged)
    for col_ini, col_fin, label in cfg["grupos"]:
        ws.cell(row=1, column=col_ini).value = label
        if col_ini < col_fin:
            ws.merge_cells(
                f"{get_column_letter(col_ini)}1:{get_column_letter(col_fin)}1"
            )

    # Fila 1 — TENDENCIA_PRECIO merged vertical (fila 1 + fila 2)
    t_letter = get_column_letter(t_col)
    ws.cell(row=1, column=t_col).value = "TENDENCIA_PRECIO"
    ws.merge_cells(f"{t_letter}1:{t_letter}2")

    # Fila 2 — etiquetas legibles (reemplaza los nombres técnicos de pandas)
    for col, label in cfg["row2_labels"].items():
        ws.cell(row=2, column=col).value = label

    # Datos — borde fino izq/der + formato numérico por columna en todas las
    # filas; la columna "Precio promedio"/"PRECIO" además en rojo negrita
    # con relleno rosado (igual que el cuadro SAS).
    p_col = cfg["precio_var_col"]
    numfmts = cfg.get("numfmts", {})
    for row in range(3, ws.max_row + 1):
        for c in range(1, t_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.border = _DATA_BORDER
            cell.alignment = _DATA_ALIGN
            if c in numfmts:
                cell.number_format = numfmts[c]
            if c == p_col:
                cell.font = _PRECIO_FONT
                cell.fill = _PRECIO_FILL

    # Formato condicional de la columna "Precio promedio"/"PRECIO": entre
    # -5% y +5% se pinta con el estilo de Excel "Relleno amarillo con texto
    # amarillo oscuro"; fuera de ese rango queda el rojo/rosado de base.
    p_letter = get_column_letter(p_col)
    cf_rule = CellIsRule(
        operator="between",
        formula=["-0.05", "0.05"],
        font=_PRECIO_CF_FONT,
        fill=_PRECIO_CF_FILL,
        stopIfTrue=True,
    )
    ws.conditional_formatting.add(f"{p_letter}3:{p_letter}{ws.max_row}", cf_rule)

    # Anchos de columna
    for col_letter, width in cfg["col_widths"].items():
        ws.column_dimensions[col_letter].width = width


# ─── API pública ──────────────────────────────────────────────────────────────

def formatear_cuadros_pub(path: Path, ant: str, act: str) -> None:
    """Aplica formato de publicación a ``CUADROS_{PERI}.xlsx`` en disco (in-place).

    Transforma el archivo escrito por pandas en el formato de entrega oficial:

    - **FINCA**: anchos de columna, encabezados sin negrita, hoja oculta.
    - **MUNICIPIO / DEPARTAMENTO / MACROREGION**: encabezados en 2 filas con
      grupos del mes anterior, mes actual, variación y tendencia (celdas
      combinadas, fuente MS Sans Serif 10pt, relleno gris claro, bordes
      finos); columna "Precio promedio" en rojo negrita con relleno rosado
      y formato de porcentaje; formato numérico específico por columna.

    Args:
        path: Ruta al archivo Excel a formatear. Debe existir y contener las
            hojas ``FINCA``, ``MUNICIPIO``, ``DEPARTAMENTO`` y ``MACROREGION``.
        ant: Iniciales del mes anterior (ej. ``'FEB'``). Se usa como etiqueta
            del grupo de columnas del mes anterior.
        act: Iniciales del mes actual (ej. ``'MAR'``). Se usa como etiqueta
            del grupo de columnas del mes actual.

    Raises:
        KeyError: Si el archivo no contiene alguna de las cuatro hojas esperadas.
        FileNotFoundError: Si ``path`` no existe.

    Example:
        >>> from pathlib import Path
        >>> formatear_cuadros_pub(Path('data/08_reporting/CUADROS_032026.xlsx'),
        ...                       ant='FEB', act='MAR')
    """
    wb = openpyxl.load_workbook(str(path))

    # FINCA: anchos + sin negrita en encabezado + ocultar hoja
    ws_finca = wb["FINCA"]
    for col_letter, width in _FINCA_WIDTHS.items():
        ws_finca.column_dimensions[col_letter].width = width
    for cell in ws_finca[1]:
        cell.font = _NOT_BOLD
    ws_finca.sheet_state = "hidden"

    # MUNICIPIO / DEPARTAMENTO / MACROREGION — encabezados en 2 filas
    _aplicar_2row(wb["MUNICIPIO"],    _cfg_municipio(ant, act))
    _aplicar_2row(wb["DEPARTAMENTO"], _cfg_departamento(ant, act))
    _aplicar_2row(wb["MACROREGION"],  _cfg_macroregion(ant, act))

    wb.save(str(path))
    wb.close()


def formatear_cuadros_tot(path: Path) -> None:
    """Aplica anchos de columna y elimina negrita en ``CUADROS_{PERI}_TOT.xlsx`` (in-place).

    El archivo TOT mantiene la estructura de encabezado en una sola fila (sin
    grupos) pero necesita anchos de columna específicos para que sea legible al
    abrir en Excel.

    Args:
        path: Ruta al archivo Excel a formatear. Debe contener las hojas
            ``FINCA``, ``MUNICIPIO``, ``DEPARTAMENTO`` y ``MACROREGION``.
            Las hojas ausentes se omiten sin error.

    Raises:
        FileNotFoundError: Si ``path`` no existe.

    Example:
        >>> from pathlib import Path
        >>> formatear_cuadros_tot(Path('data/08_reporting/CUADROS_032026_TOT.xlsx'))
    """
    wb = openpyxl.load_workbook(str(path))

    for sheet_name, widths in _TOT_WIDTHS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.font = _NOT_BOLD
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

    wb.save(str(path))
    wb.close()
