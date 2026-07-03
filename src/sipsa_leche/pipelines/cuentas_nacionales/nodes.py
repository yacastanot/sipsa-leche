"""Nodos del pipeline cuentas_nacionales — M11: Reportes de Leche Cruda para DSCN.

Genera dos archivos de salida cada mes:

  ``LECHE_CRUDA_{PERI}.xlsx``
      Copia del histórico con la fila del mes calculada internamente:
      D = producción macro recolectada,  E = factor corrección,
      F = D × E,  G = variación mensual,  H = variación anual,  J = índice acumulado.

  ``Excluidas_leche_{PERI}.xlsx``
      Copia del panel histórico de fincas atípicas con la columna del mes
      poblada desde ``variacion_finca``; Hoja2 reconstruida con el dump del
      panel completo (IDFINCA, T_PROD_{ant}, T_PROD_{act}).

Diagrama de dependencias:
    semanas_validacion                                 → calcular_semanas_operativo → n_semanas_operativo
    variacion_finca + ruta_excluidas + params          → calcular_excluidas         → total_excluidas_mes
    variacion_macro + total_excluidas_mes + n_semanas  → generar_leche_cruda        → leche_cruda_resumen
"""
from __future__ import annotations

import calendar
import re
import shutil
from copy import copy
from pathlib import Path

import openpyxl
import pandas as pd
import structlog

log = structlog.get_logger()


# ─── Helpers privados ─────────────────────────────────────────────────────────

def _col_prodc(periodo: str) -> str:
    """Nombre de columna PRODC{MM}{YYYY} para el periodo MMAAAA dado."""
    return f"PRODC{periodo[:2]}{periodo[2:]}"


def _dias_mes(periodo: str) -> int:
    """Número de días del mes derivado del código MMAAAA."""
    return calendar.monthrange(int(periodo[2:]), int(periodo[:2]))[1]


def _trimestre(periodo: str) -> str:
    """Trimestre (I/II/III/IV) del mes en el periodo MMAAAA."""
    mes = int(periodo[:2])
    return ["I", "I", "I", "II", "II", "II", "III", "III", "III", "IV", "IV", "IV"][mes - 1]


def _copia_estilo(src: openpyxl.cell.Cell, dst: openpyxl.cell.Cell) -> None:
    """Copia el estilo (fuente, relleno, borde, alineación, formato) de src a dst."""
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format


def _float_safe(val) -> float | None:
    """Convierte val a float; devuelve None si es None, '#REF!' u otro error."""
    if val is None:
        return None
    try:
        result = float(val)
        return result
    except (TypeError, ValueError):
        return None


def _is_empty_value(val) -> bool:
    """Identifica celdas vacias o NaN en datos provenientes de Excel/parquet."""
    if val is None:
        return True
    if isinstance(val, str):
        return val.strip() == ""
    try:
        return bool(pd.isna(val))
    except TypeError:
        return False


_SAFE_FORMULA = re.compile(r"^[\d\s\+\-\*\/\(\)\.]+$")


def _eval_formula(val) -> float | None:
    """Evalúa fórmulas aritméticas simples sin referencias a celdas.

    Acepta valores directos (int/float) y strings con fórmula del tipo
    ``'=20482264-1818614'`` o ``'=28/(7*4)'``. Devuelve None para
    fórmulas con referencias a celdas (ej. ``'=D172*E172'``).
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s.startswith("="):
        s = s[1:]
    if _SAFE_FORMULA.match(s):
        try:
            return float(eval(s))  # solo aritmética pura, sin referencias
        except Exception:
            return None
    return None  # fórmula con referencias a celdas — se calcula en cadena


def _reconstruir_historico_leche(ws) -> dict[int, dict[str, float]]:
    """Reconstruye la serie histórica F, G, J a partir de las fórmulas del libro.

    Las filas 2-15 del archivo histórico no tienen D ni E (solo J como número
    directo); la cadena se construye desde la primera fila con ambos valores.
    D y E son aritmética pura (sin referencias a celdas) y se evalúan con
    ``_eval_formula``. F, G, J se encadenan fila a fila hacia adelante.

    Args:
        ws: Hoja de cálculo ``'LECHE CRUDA'`` abierta con ``data_only=False``.

    Returns:
        Diccionario ``{fila: {'F': float, 'G': float | None, 'J': float | None}}``.
    """
    DATA_START = 3
    result: dict[int, dict] = {}
    chain_started = False

    for r in range(DATA_START, ws.max_row + 1):
        anno = ws.cell(r, 1).value
        if anno is None:
            if chain_started:
                break  # fin real de los datos
            continue

        d_val = _eval_formula(ws.cell(r, 4).value)
        e_val = _eval_formula(ws.cell(r, 5).value)
        j_raw = ws.cell(r, 10).value

        if d_val is None or e_val is None:
            if chain_started:
                break  # D/E vacíos después del inicio → fin de cadena
            continue  # filas pre-cadena 2011-2012 sin D/E

        chain_started = True
        F_curr = d_val * e_val
        F_prev = result[r - 1]["F"] if r - 1 in result else None
        G_curr = (F_curr / F_prev - 1) if F_prev else None

        # J: valor numérico directo (filas âncora) o fórmula encadenada =(G+1)*J_prev
        j_direct = _eval_formula(j_raw)
        if j_direct is not None and not str(j_raw or "").startswith("="):
            J_curr = j_direct
        elif G_curr is not None and r - 1 in result and result[r - 1]["J"] is not None:
            J_curr = (G_curr + 1) * result[r - 1]["J"]
        else:
            J_curr = None

        result[r] = {"F": F_curr, "G": G_curr, "J": J_curr}

    return result


# ─── Node 1: semanas ──────────────────────────────────────────────────────────

def calcular_semanas_operativo(semanas_validacion: pd.DataFrame) -> int:
    """Cuenta las semanas del operativo a partir del CSV de semanas.

    Args:
        semanas_validacion: DataFrame con una fila por semana. Columnas:
            ``periodo``, ``mes``, ``semana``, ``fincas``, ``registros``.

    Returns:
        Número de semanas distintas del operativo (típicamente 4 ó 5).
    """
    n = len(semanas_validacion)
    log.info("semanas_operativo", n_semanas=n)
    return n


# ─── Node 2: excluidas ────────────────────────────────────────────────────────

def calcular_excluidas(
    variacion_finca: pd.DataFrame,
    ruta_excluidas: str,
    periodo: str,
    mes_actual: str,
    mes_anterior: str,
) -> float:
    """Actualiza ``Excluidas_leche_{PERI}.xlsx`` y devuelve la producción excluida.

    Lee Hoja1 del archivo base para obtener las fincas atípicas, cruza con
    ``variacion_finca`` para la producción del mes actual, agrega la columna
    ``PRODC{MM}{YYYY}`` y reconstruye Hoja2 con el panel completo del periodo.
    Guarda el resultado en ``data/08_reporting/Excluidas_leche_{PERI}.xlsx``.

    Args:
        variacion_finca: DataFrame con columnas ``IDFINCA``,
            ``T_PROD_{mes_anterior}`` y ``T_PROD_{mes_actual}``.
        ruta_excluidas: Ruta al archivo ``Excluidas_leche.xlsx`` base.
        periodo: Código MMAAAA (ej. ``'032026'``).
        mes_actual: Iniciales del mes actual (ej. ``'MAR'``).
        mes_anterior: Iniciales del mes anterior (ej. ``'FEB'``).

    Returns:
        Producción total (litros) de las fincas excluidas en el mes actual.
    """
    ruta_base = Path(ruta_excluidas)
    col_act = f"T_PROD_{mes_actual}"
    col_ant = f"T_PROD_{mes_anterior}"
    col_nueva = _col_prodc(periodo)

    # ── Mapa de producción desde variacion_finca ───────────────────────────
    finca_df = variacion_finca.copy()
    finca_df["_IDF"] = finca_df["IDFINCA"].astype(str).str.strip().str.zfill(7)
    prod_act_map: dict[str, float | None] = finca_df.set_index("_IDF")[col_act].to_dict()

    # ── Abrir para lectura/escritura ───────────────────────────────────────
    wb = openpyxl.load_workbook(str(ruta_base), data_only=False)
    ws1 = wb["Hoja1"]

    # Hoja1 tiene la fila de encabezados en la fila 2 (fila 1 está vacía)
    HEADER_ROW = 2
    DATA_START  = 3
    headers1 = [ws1.cell(HEADER_ROW, c).value for c in range(1, ws1.max_column + 1)]
    col_idf_idx  = headers1.index("IDFINCA") + 1   # 1-based
    col_dept_idx = 1                                 # columna A = DEPARTAMENTO

    # Clasificar filas: fincas excluidas / total / variacion
    excluded_idf: list[str] = []
    excluded_rows: list[int] = []
    row_total: int | None = None
    row_variacion: int | None = None
    row_salen: int | None = None
    row_entran: int | None = None
    row_diferencia: int | None = None

    for r in range(DATA_START, ws1.max_row + 1):
        dept = ws1.cell(r, col_dept_idx).value
        idf  = ws1.cell(r, col_idf_idx).value
        marca = str(dept).strip().lower() if dept is not None else ""
        concepto = ws1.cell(r, 3).value
        concepto_norm = str(concepto).strip().lower() if concepto is not None else ""

        if marca == "total":
            row_total = r
        elif marca == "variacion":
            row_variacion = r
        elif concepto_norm == "salen":
            row_salen = r
        elif concepto_norm == "entran":
            row_entran = r
        elif concepto_norm == "diferencia":
            row_diferencia = r
        elif idf and str(idf).strip():
            excluded_idf.append(str(idf).strip().zfill(7))
            excluded_rows.append(r)

    # ── Calcular producción por finca excluida ─────────────────────────────
    # prod_act_map.get() puede devolver NaN (finca presente en variacion_finca
    # pero sin producción ese mes) además de None (finca ausente del todo).
    # Ambos casos se tratan como "sin dato" — pd.notna() filtra los dos.
    prods: list[float | None] = [prod_act_map.get(idf) for idf in excluded_idf]
    total_excluidas = sum(p for p in prods if pd.notna(p))
    ant_empty = finca_df[col_ant].map(_is_empty_value)
    act_empty = finca_df[col_act].map(_is_empty_value)
    total_entran = pd.to_numeric(finca_df.loc[ant_empty, col_act], errors="coerce").sum()
    total_salen = pd.to_numeric(finca_df.loc[act_empty, col_ant], errors="coerce").sum()

    n_encontradas = sum(1 for p in prods if pd.notna(p))
    log.info(
        "excluidas_produccion",
        col_nueva=col_nueva,
        n_fincas=len(excluded_idf),
        n_encontradas=n_encontradas,
        n_sin_datos=len(excluded_idf) - n_encontradas,
        total_excluidas=int(total_excluidas),
        total_salen=int(total_salen),
        total_entran=int(total_entran),
    )

    # ── Determinar índice de columna nueva en Hoja1 ────────────────────────
    matching_cols = [
        idx + 1
        for idx, header in enumerate(headers1)
        if header is not None and str(header).strip() == col_nueva
    ]
    if matching_cols:
        c_nueva_idx = matching_cols[-1]
        if len(matching_cols) > 1:
            log.warning(
                "columna_periodo_duplicada",
                col_nueva=col_nueva,
                columnas=matching_cols,
            )
    else:
        c_nueva_idx = ws1.max_column + 1
        h_cell = ws1.cell(HEADER_ROW, c_nueva_idx)
        h_cell.value = col_nueva
        _copia_estilo(ws1.cell(HEADER_ROW, c_nueva_idx - 1), h_cell)
        prev_col_w = openpyxl.utils.get_column_letter(c_nueva_idx - 1)
        nueva_col_w = openpyxl.utils.get_column_letter(c_nueva_idx)
        if prev_col_w in ws1.column_dimensions:
            ws1.column_dimensions[nueva_col_w].width = ws1.column_dimensions[prev_col_w].width

    # ── Escribir producción por finca ──────────────────────────────────────
    for r, prod in zip(excluded_rows, prods, strict=True):
        cell = ws1.cell(r, c_nueva_idx)
        cell.value = int(prod) if pd.notna(prod) else None
        _copia_estilo(ws1.cell(r, c_nueva_idx - 1), cell)

    # ── Actualizar fila total ──────────────────────────────────────────────
    if row_total is not None:
        cell = ws1.cell(row_total, c_nueva_idx)
        cell.value = int(total_excluidas)
        _copia_estilo(ws1.cell(row_total, c_nueva_idx - 1), cell)

    # ── Actualizar fila de referencia limpia (total + 3) ──────────────────
    # Esta fila (sin marcador en col A) almacena el total del período como
    # valor de referencia directo para el cálculo de D en LECHE_CRUDA.
    if row_total is not None:
        row_ref = row_total + 3
        cell = ws1.cell(row_ref, c_nueva_idx)
        cell.value = int(total_excluidas)
        _copia_estilo(ws1.cell(row_ref, c_nueva_idx - 1), cell)

    # ── Actualizar fila variación como fórmula ────────────────────────────
    col_letra      = openpyxl.utils.get_column_letter(c_nueva_idx)
    prev_col_letra = openpyxl.utils.get_column_letter(c_nueva_idx - 1)
    if row_total is not None and row_variacion is not None:
        cell = ws1.cell(row_variacion, c_nueva_idx)
        cell.value = f"=({col_letra}{row_total}-{prev_col_letra}{row_total})/{prev_col_letra}{row_total}"
        _copia_estilo(ws1.cell(row_variacion, c_nueva_idx - 1), cell)

    # ── Actualizar salidas/entradas del panel ──────────────────────────────
    if row_salen is not None:
        _copia_estilo(ws1.cell(row_salen, c_nueva_idx - 1), ws1.cell(row_salen, c_nueva_idx))
        ws1.cell(row_salen, c_nueva_idx).value = int(total_salen)
    if row_entran is not None:
        _copia_estilo(ws1.cell(row_entran, c_nueva_idx - 1), ws1.cell(row_entran, c_nueva_idx))
        ws1.cell(row_entran, c_nueva_idx).value = int(total_entran)
    if row_diferencia is not None and row_entran is not None and row_salen is not None:
        _copia_estilo(
            ws1.cell(row_diferencia, c_nueva_idx - 1),
            ws1.cell(row_diferencia, c_nueva_idx),
        )
        ws1.cell(row_diferencia, c_nueva_idx).value = (
            f"=+{col_letra}{row_entran}-{col_letra}{row_salen}"
        )

    # ── Reconstruir Hoja2: dump del panel completo ─────────────────────────
    ws2 = wb["Hoja2"]
    # Limpiar todas las filas de datos (conservar estructura de columnas)
    for r in range(ws2.max_row, 1, -1):
        ws2.delete_rows(r)

    ws2.cell(1, 1).value = "IDFINCA"
    ws2.cell(1, 2).value = col_ant   # ej. T_PROD_FEB
    ws2.cell(1, 3).value = col_act   # ej. T_PROD_MAR

    panel_df = finca_df[["IDFINCA", col_ant, col_act]].copy()
    for r_idx, (_, row) in enumerate(panel_df.iterrows(), start=2):
        ws2.cell(r_idx, 1).value = str(row["IDFINCA"])
        ant_raw = row[col_ant]
        act_raw = row[col_act]
        ws2.cell(r_idx, 2).value = int(ant_raw) if (pd.notna(ant_raw) and ant_raw is not None) else None
        ws2.cell(r_idx, 3).value = int(act_raw) if (pd.notna(act_raw) and act_raw is not None) else None

    # ── Guardar en data/08_reporting ───────────────────────────────────────
    output_path = Path("data/08_reporting") / f"Excluidas_leche_{periodo}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    wb.close()
    log.info("excluidas_guardadas", ruta=str(output_path))

    # ── Promover como nuevo archivo base para el siguiente período ─────────
    base_dst = Path(ruta_excluidas)
    shutil.copy2(str(output_path), str(base_dst))
    log.info("excluidas_base_actualizada", ruta=str(base_dst))

    return float(total_excluidas)


# ─── Node 3: generar LECHE_CRUDA ──────────────────────────────────────────────

_MES_NUM = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}
_MES_NOMBRE = {
    1: "enero",      2: "febrero",    3: "marzo",
    4: "abril",      5: "mayo",       6: "junio",
    7: "julio",      8: "agosto",     9: "septiembre",
    10: "octubre",   11: "noviembre", 12: "diciembre",
}
_TRIM_MESES = {"I": [1,2,3], "II": [4,5,6], "III": [7,8,9], "IV": [10,11,12]}


def generar_leche_cruda(
    variacion_macro: pd.DataFrame,
    total_excluidas: float,
    n_semanas: int,
    ruta_leche_cruda_base: str,
    periodo: str,
    mes_actual: str,
) -> dict:
    """Genera ``LECHE_CRUDA_{PERI}.xlsx`` con la fila del mes calculada.

    Escribe todos los valores calculados internamente (sin fórmulas Excel):

    - LECHE CRUDA · D = Σ(T_PRODUCCION_MACRO{act}), E = días/(7×semanas),
      F = D×E, G = F/F_ant−1, H = (F−F_12ago)/F_12ago×100, J = (G+1)×J_ant
    - LecheDANE · D = F de LECHE CRUDA (pegado como valor),
      E = variación mensual, F = variación anual
    - trimes · D = suma de LecheDANE del trimestre, E/F = variaciones trimestrales

    Args:
        variacion_macro: DataFrame con columna ``T_PRODUCCION_MACRO{mes_actual}``.
        total_excluidas: Suma de producción de fincas excluidas (litros).
            Se conserva para trazabilidad y para forzar la generación previa
            de ``Excluidas_leche_{PERI}.xlsx``; no se descuenta en columna D.
        n_semanas: Número de semanas del operativo del mes.
        ruta_leche_cruda_base: Ruta al archivo ``LECHE_CRUDA_EST_BASE.xlsx``.
        periodo: Código MMAAAA (ej. ``'042026'``).
        mes_actual: Iniciales del mes (ej. ``'ABR'``).

    Returns:
        Diccionario con los valores calculados para trazabilidad.
    """
    mes_nombre = _MES_NOMBRE[int(periodo[:2])]
    # ── Cálculos fundamentales ────────────────────────────────────────────
    col_macro = f"T_PRODUCCION_MACRO{mes_actual}"
    total_macro = float(variacion_macro[col_macro].sum())
    dias  = _dias_mes(periodo)
    anio  = int(periodo[2:])
    E = dias / (7 * n_semanas)

    ruta_base = Path(ruta_leche_cruda_base)

    # ── Reconstruir histórico de F, G, J desde el archivo base ───────────
    wb_hist = openpyxl.load_workbook(str(ruta_base), data_only=False)
    ws_hist = wb_hist["LECHE CRUDA"]
    historico = _reconstruir_historico_leche(ws_hist)
    wb_hist.close()

    if not historico:
        raise ValueError("No se encontraron datos históricos en 'LECHE CRUDA' columna F")

    # ── Copiar base y abrir para escritura ────────────────────────────────
    output_path = Path("data/08_reporting") / f"LECHE_CRUDA_{periodo}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(str(ruta_base), str(output_path))

    wb_w = openpyxl.load_workbook(str(output_path), data_only=False)
    ws_w = wb_w["LECHE CRUDA"]

    # Detectar fila del periodo (sobreescribir si ya existe, añadir si no)
    last_data_row = max(historico.keys())
    new_row = last_data_row + 1
    for r in range(2, ws_w.max_row + 1):
        anno_cell = ws_w.cell(r, 1).value
        mes_cell  = ws_w.cell(r, 2).value
        if (anno_cell == anio
                and mes_cell is not None
                and mes_nombre.lower() in str(mes_cell).lower()):
            new_row = r
            log.info("sobreescribiendo_fila_existente", row=r)
            break

    prev_row    = new_row - 1
    row_12_ago  = new_row - 12

    F_prev   = historico.get(prev_row, {}).get("F")
    J_prev   = historico.get(prev_row, {}).get("J")
    F_12_ago = historico.get(row_12_ago, {}).get("F") if row_12_ago >= 3 else None

    if F_prev is None:
        raise ValueError(
            f"No se pudo reconstruir F en fila {prev_row} del histórico"
        )

    # ── Calcular D = producción recolectada − excluidas (fila 37 Excluidas) ──
    D = total_macro - total_excluidas

    # ── Calcular F, G, H, J ───────────────────────────────────────────────
    F = D * E
    G = F / F_prev - 1
    H = (F - F_12_ago) / F_12_ago * 100 if F_12_ago else None
    J = (G + 1) * J_prev if J_prev is not None else None

    log.info(
        "leche_cruda_calculada",
        anio=anio, mes=mes_nombre, fila=new_row,
        total_macro=int(total_macro), total_excluidas=int(total_excluidas),
        D=int(D), E=round(E, 6),
        F=round(F, 2), G=round(G, 6),
        H=round(H, 4) if H else None,
        J=round(J, 2) if J else None,
    )

    # ── Escribir hoja LECHE CRUDA (todos como valores) ────────────────────
    def _wv(col: int, val, nf: str | None = None) -> None:
        cell = ws_w.cell(new_row, col)
        cell.value = val
        _copia_estilo(ws_w.cell(prev_row, col), cell)
        if nf:
            cell.number_format = nf

    _wv(1, anio)
    _wv(2, mes_nombre)
    _wv(3, _trimestre(periodo))
    _wv(4, f"={int(total_macro)}-{int(total_excluidas)}", "#,##0.00")
    _wv(5, f"={dias}/(7*{n_semanas})",     "0.00")
    _wv(6, f"=D{new_row}*E{new_row}",      "#,##0.00")
    _wv(7, f"=(F{new_row}/F{prev_row})-1", "0.00%")
    if row_12_ago >= 3:
        _wv(8, f"=((F{new_row}-F{row_12_ago})/F{row_12_ago})*100")
    _wv(10, f"=(G{new_row}+1)*J{prev_row}", "#,##0")
    if ws_w.row_dimensions[prev_row].height:
        ws_w.row_dimensions[new_row].height = ws_w.row_dimensions[prev_row].height

    # ── Actualizar hoja LecheDANE: D como referencia cruzada, E y F como fórmulas ──
    ws_ld = wb_w["LecheDANE"]

    ld_row: int | None = None
    for r in range(2, ws_ld.max_row + 1):
        if (ws_ld.cell(r, 1).value == anio
                and ws_ld.cell(r, 2).value is not None
                and mes_nombre.lower() in str(ws_ld.cell(r, 2).value).lower()):
            ld_row = r
            break
    if ld_row is None:
        ld_row = ws_ld.max_row + 1

    ld_prev_row   = ld_row - 1
    ld_12_ago_row = ld_row - 12

    # Leer valores anteriores ANTES de sobreescribir (para comparación y suma trimestral)
    base_ld_d   = _float_safe(ws_ld.cell(ld_row, 4).value)
    D_ld_prev   = _float_safe(ws_ld.cell(ld_prev_row, 4).value)
    D_ld_12_ago = _float_safe(ws_ld.cell(ld_12_ago_row, 4).value) if ld_12_ago_row >= 2 else None
    F_rounded   = round(F, 2)
    E_ld        = (F_rounded / D_ld_prev   - 1) if D_ld_prev   else None
    F_ld        = (F_rounded / D_ld_12_ago - 1) if D_ld_12_ago else None

    for col, val in [(1, anio), (2, mes_nombre), (3, _trimestre(periodo))]:
        cell = ws_ld.cell(ld_row, col)
        cell.value = val
        _copia_estilo(ws_ld.cell(ld_prev_row, col), cell)

    for col, formula, nf in [
        (4, f"='LECHE CRUDA'!F{new_row}",           "#,##0.00"),
        (5, f"=(D{ld_row}/D{ld_prev_row})-1",       "0.00%"),
        (6, f"=(D{ld_row}/D{ld_12_ago_row})-1",     "0.00%"),
    ]:
        cell = ws_ld.cell(ld_row, col)
        cell.value = formula
        _copia_estilo(ws_ld.cell(ld_prev_row, col), cell)
        if nf:
            cell.number_format = nf
    if ws_ld.row_dimensions[ld_prev_row].height:
        ws_ld.row_dimensions[ld_row].height = ws_ld.row_dimensions[ld_prev_row].height

    log.info(
        "lecheDane_actualizada",
        fila=ld_row,
        D_formula=f"='LECHE CRUDA'!F{new_row}",
        D_valor_calculado=F_rounded,
        E_variacion_mensual=round(E_ld, 6) if E_ld is not None else None,
        F_variacion_anual=round(F_ld, 6) if F_ld is not None else None,
    )

    # ── Comparación contra el archivo base ────────────────────────────────
    if base_ld_d is not None:
        log.info(
            "comparacion_base",
            LecheDANE_D_pipeline=F_rounded,
            LecheDANE_D_base=round(base_ld_d, 2),
            diferencia=round(F_rounded - base_ld_d, 2),
        )

    if hasattr(wb_w, "calculation"):
        wb_w.calculation.fullCalcOnLoad = True
        wb_w.calculation.forceFullCalc = True

    wb_w.save(str(output_path))
    wb_w.close()
    log.info("leche_cruda_guardada", ruta=str(output_path))

    # ── Promover como nuevo archivo base para el siguiente período ─────────
    base_dst = Path(ruta_leche_cruda_base)
    shutil.copy2(str(output_path), str(base_dst))
    log.info("leche_cruda_base_actualizada", ruta=str(base_dst))

    return {
        "total_macro":              int(total_macro),
        "D_produccion_recolectada": int(D),
        "E_factor_correccion":      round(E, 6),
        "F_producido_DANE":         round(F, 2),
        "G_variacion_mensual":      round(G, 6),
        "H_variacion_anual":        round(H, 4) if H is not None else None,
        "J_indice_acumulado":       round(J, 2) if J is not None else None,
        "LecheDANE_D":              F_rounded,
        "LecheDANE_E":              round(E_ld, 6) if E_ld is not None else None,
        "LecheDANE_F":              round(F_ld, 6) if F_ld is not None else None,
        "n_semanas":                n_semanas,
        "total_excluidas":          int(total_excluidas),
        "fila_LECHE_CRUDA":         new_row,
        "fila_LecheDANE":           ld_row,
    }
